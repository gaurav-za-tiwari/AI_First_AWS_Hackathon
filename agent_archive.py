"""
agent_archive.py
================
Agent 4 — ArchiveAgent

Runs after ReportAgent. Two archival responsibilities:

1. ARCHIVE HISTORICAL — move "Historical Alerts/alerts_data.xlsx" into
   "KB_Processed/" with a UTC timestamp prefix, then merge it into the
   cumulative master file (KB_Processed/master_kb.xlsx).
   After the move, "Historical Alerts/" is empty and ready for the next
   batch of historical data.

2. ARCHIVE OPEN ALERTS — move "Open Alerts/open_alerts_data.xlsx" into
   "Processed_Alert/" with the same UTC timestamp prefix.
   The file is only moved after the closure report has been successfully
   written, so it is never lost if the pipeline fails mid-run.

Note: The dated closure report in "Closure Report/" is the final output
      of each run and is never moved or deleted by this agent.
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config as cfg
from orchestrator import BaseAgent, Task

log = logging.getLogger("AML-A2A")


class ArchiveAgent(BaseAgent):
    name = "ArchiveAgent"

    # ── Public entry point ─────────────────────────────────────────────────────

    async def handle(self, task: Task) -> dict:
        historical         = Path(task.input["historical_file"])
        open_alerts        = Path(task.input["open_alerts_file"])
        report_file        = Path(task.input["report_file"])
        kb_dir             = Path(task.input["kb_processed_dir"])
        processed_alert_dir= Path(task.input["processed_alert_dir"])
        run_ts             = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

        # ── Validate inputs ────────────────────────────────────────────────────
        if not historical.exists():
            raise FileNotFoundError(
                f"[ArchiveAgent] Historical file not found: {historical}\n"
                f"  Place alerts_data.xlsx inside the 'Historical Alerts/' folder."
            )
        if not open_alerts.exists():
            raise FileNotFoundError(
                f"[ArchiveAgent] Open alerts file not found: {open_alerts}\n"
                f"  Place open_alerts_data.xlsx inside the 'Open Alerts/' folder."
            )
        if not report_file.exists():
            raise FileNotFoundError(
                f"[ArchiveAgent] Closure report not found: {report_file}\n"
                f"  ReportAgent must complete successfully before ArchiveAgent runs."
            )

        # ── Step 1 — ensure archive directories exist ──────────────────────────
        kb_dir.mkdir(parents=True, exist_ok=True)
        processed_alert_dir.mkdir(parents=True, exist_ok=True)

        # ── Step 2 — MOVE historical file → KB_Processed/<ts>_alerts_data.xlsx
        # Move (not copy) so "Historical Alerts/" is empty after this run and
        # ready for the next batch of historical data.
        archived_hist_name = f"{run_ts}_alerts_data.xlsx"
        archived_hist_path = kb_dir / archived_hist_name
        shutil.move(str(historical), str(archived_hist_path))
        log.info("[ArchiveAgent] Historical: moved %s → KB_Processed/%s",
                 historical.name, archived_hist_name)

        # ── Step 3 — merge historical snapshot into master_kb.xlsx ────────────
        master_path  = kb_dir / "master_kb.xlsx"
        merge_result = self._merge_into_master(archived_hist_path, master_path, run_ts)

        # ── Step 4 — MOVE open alerts → Processed_Alert/<ts>_open_alerts_data.xlsx
        # Only moved after the report is confirmed to exist (validated above),
        # so the source file is never lost if the pipeline failed earlier.
        archived_open_name = f"{run_ts}_open_alerts_data.xlsx"
        archived_open_path = processed_alert_dir / archived_open_name
        shutil.move(str(open_alerts), str(archived_open_path))
        log.info("[ArchiveAgent] Open alerts: moved %s → Processed_Alert/%s",
                 open_alerts.name, archived_open_name)

        log.info("[ArchiveAgent] Historical Alerts/ and Open Alerts/ are now empty "
                 "— drop fresh files there for the next run.")

        return {
            "archived_historical":    archived_hist_name,
            "archived_open_alerts":   archived_open_name,
            "master_kb":              str(master_path),
            "master_total_rows":      merge_result["total_rows"],
            "master_run_count":       merge_result["run_count"],
        }

    # ── Merge logic ────────────────────────────────────────────────────────────

    def _merge_into_master(self, source: Path, master: Path,
                           run_ts: str) -> dict:
        """
        Reads the just-archived snapshot and appends every sheet's rows into
        the matching sheet of master_kb.xlsx.

        - First run: creates master_kb.xlsx from scratch.
        - Subsequent runs: loads existing master and concatenates new rows.
        - Sheet names are matched by name; new sheets are created if needed.
        - A _run_timestamp column is stamped on every incoming row.
        """
        source_sheets: dict[str, pd.DataFrame] = pd.read_excel(
            source, sheet_name=None
        )
        for sheet_name, df in source_sheets.items():
            df["_run_timestamp"] = run_ts
            source_sheets[sheet_name] = df

        if not master.exists():
            log.info("[ArchiveAgent] master_kb.xlsx not found — creating fresh.")
            with pd.ExcelWriter(str(master), engine="openpyxl") as writer:
                for sheet_name, df in source_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            total_rows = sum(len(df) for df in source_sheets.values())
            self._apply_master_formatting(master)
            log.info("[ArchiveAgent] master_kb.xlsx created with %d rows.", total_rows)
            return {"total_rows": total_rows, "run_count": 1}

        # Subsequent runs — load and concatenate
        existing_sheets: dict[str, pd.DataFrame] = pd.read_excel(
            master, sheet_name=None
        )
        sample_df = next(iter(existing_sheets.values()), pd.DataFrame())
        run_count = (sample_df["_run_timestamp"].nunique() + 1
                     if "_run_timestamp" in sample_df.columns else 2)

        merged_sheets: dict[str, pd.DataFrame] = {}
        for sheet_name in set(existing_sheets) | set(source_sheets):
            merged_sheets[sheet_name] = pd.concat(
                [existing_sheets.get(sheet_name, pd.DataFrame()),
                 source_sheets.get(sheet_name, pd.DataFrame())],
                ignore_index=True, sort=False,
            )

        total_rows = sum(len(df) for df in merged_sheets.values())
        with pd.ExcelWriter(str(master), engine="openpyxl") as writer:
            for sheet_name, df in merged_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        self._apply_master_formatting(master)
        log.info("[ArchiveAgent] master_kb.xlsx updated — run #%d, %d rows.",
                 run_count, total_rows)
        return {"total_rows": total_rows, "run_count": run_count}

    # ── Formatting ─────────────────────────────────────────────────────────────

    def _apply_master_formatting(self, master: Path):
        """Freeze header row, bold navy headers, highlight _run_timestamp column."""
        wb = load_workbook(str(master))
        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue
            ws.freeze_panes = "A2"
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

            for cell in ws[1]:
                cell.font      = Font(bold=True, name="Calibri", size=10,
                                      color="FFFFFF")
                cell.fill      = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Highlight _run_timestamp column in muted teal for easy identification
            for col_idx, cell in enumerate(ws[1], 1):
                if str(cell.value) == "_run_timestamp":
                    for row_idx in range(1, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            "solid", start_color="D6E4F0"
                        )
        wb.save(str(master))
