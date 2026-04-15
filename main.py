"""
AML Alert Auto-Closure System — Google A2A Agent Architecture
=============================================================
Agent 1 (EmbeddingAgent) : Reads alerts_data.xlsx, embeds each alert row into
                           a vector and upserts into ChromaDB (persisted on disk).
                           Skips rows already embedded on re-runs — incremental only.
Agent 2 (AnalysisAgent)  : For each open alert, embeds it then queries ChromaDB for
                           the top-K most similar historical alerts. Only that small
                           context window is sent to the LLM — no full-file dumps.
Agent 3 (ReportAgent)    : Produces an auto-closure report and writes decisions +
                           comments back into open_alerts_data.xlsx.
Agent 4 (ArchiveAgent)   : Archives alerts_data.xlsx into KB_Processed/ (merging
                           every run into one cumulative master file), then promotes
                           auto_closure_report.xlsx → alerts_data.xlsx so the next
                           pipeline run trains on fresh, enriched data.
"""

import asyncio
import json
import logging
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# RAG dependencies — install with:
#   pip install chromadb sentence-transformers
#
# chromadb              : local vector database, persisted to disk — no server needed.
# sentence-transformers : runs embedding models entirely inside this Python process.
#                         No daemon, no extra server, no API key.
#                         Models are cached in ~/.cache/huggingface after first download.
#
# Embedding model: all-MiniLM-L6-v2
#   - 22 MB on disk, 384-dimensional vectors
#   - Runs on CPU in ~50 ms per alert
#   - Downloaded once from Hugging Face on first use, then works fully offline
try:
    import chromadb
    from chromadb.config import Settings
    from sentence_transformers import SentenceTransformer
    RAG_AVAILABLE = True
except ImportError as _rag_err:
    RAG_AVAILABLE = False
    _rag_import_err = str(_rag_err)

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  [%(name)s]  %(message)s",
)
log = logging.getLogger("AML-A2A")

# ── Config ─────────────────────────────────────────────────────────────────────
LLM_ENDPOINT = "http://wiphackq0vcsii.cloudloka.com:8000/v1/completions"
LLM_MODEL = "meta-llama/Meta-Llama-3.1-8B-Instruct"
MAX_TOKENS = 512
TEMPERATURE = 0.3          # lower = more deterministic / consistent

# Timeout tuple: (connect_timeout_seconds, read_timeout_seconds)
# connect — how long to wait for the TCP handshake to complete
# read    — how long to wait for the model to stream back the full response;
#           large prompts with 500+ tokens can take 90-120 s on small GPUs
LLM_TIMEOUT_CONNECT = 15    # seconds to establish connection
LLM_TIMEOUT_READ    = 180   # seconds to receive the full completion
LLM_RETRIES         = 2     # number of retry attempts on timeout / 5xx

# All paths resolve relative to THIS script's directory.
# server.py lives in the same folder and serves open_alerts_data.xlsx
# automatically — Agent 3 writes here so the browser picks it up instantly.
_HERE            = Path(__file__).parent.resolve()
HISTORICAL_FILE  = str(_HERE / "alerts_data.xlsx")
OPEN_ALERTS_FILE = str(_HERE / "open_alerts_data.xlsx")
OUTPUT_FILE      = str(_HERE / "auto_closure_report.xlsx")
KB_PROCESSED_DIR = _HERE / "KB_Processed"   # archive folder — created on first run

# ── RAG / Vector DB config ─────────────────────────────────────────────────────
VECTOR_DB_DIR    = str(_HERE / "vector_db")   # ChromaDB persists here
COLLECTION_NAME  = "aml_alerts"               # one collection, all sheets

# sentence-transformers embedding model.
# Model comparison (all run locally after first download):
#   all-MiniLM-L6-v2        → 384-dim, ~22 MB,   ~50 ms/alert  ← default
#   all-mpnet-base-v2        → 768-dim, ~420 MB,  ~120 ms/alert (higher accuracy)
#   paraphrase-MiniLM-L3-v2 → 384-dim, ~17 MB,   ~25 ms/alert  (fastest)
EMBED_MODEL_NAME = "all-MiniLM-L6-v2"

# Module-level singleton so the model is loaded exactly once per process,
# not once per alert call.
_EMBED_MODEL: "SentenceTransformer | None" = None

RAG_TOP_K       = 5    # similar historical alerts retrieved per open alert
RAG_FALLBACK_KB = True # fall back to rule-based decisions if RAG unavailable


# ══════════════════════════════════════════════════════════════════════════════
# A2A PROTOCOL PRIMITIVES
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Task:
    """Minimal A2A Task envelope."""
    id: str = field(default_factory=lambda: str(uuid.uuid4()))
    name: str = ""
    input: dict = field(default_factory=dict)
    output: dict = field(default_factory=dict)
    status: str = "pending"   # pending | running | done | error
    created_at: str = field(default_factory=lambda: datetime.utcnow().isoformat())


@dataclass
class Message:
    """A2A inter-agent message."""
    sender: str
    recipient: str
    task_id: str
    payload: dict = field(default_factory=dict)


class A2AOrchestrator:
    """
    Lightweight Google A2A-style orchestrator.
    Routes tasks between registered agents and maintains a shared
    artifact store (knowledge_base) that agents can read/write.
    """

    def __init__(self):
        self._agents: dict[str, "BaseAgent"] = {}
        self.knowledge_base: dict[str, Any] = {}
        self.task_log: list[Task] = []

    def register(self, agent: "BaseAgent"):
        self._agents[agent.name] = agent
        agent.orchestrator = self
        log.info("Registered agent: %s", agent.name)

    async def send(self, msg: Message) -> Task:
        agent = self._agents.get(msg.recipient)
        if not agent:
            raise ValueError(f"Unknown agent: {msg.recipient}")
        task = Task(id=msg.task_id, name=msg.recipient, input=msg.payload)
        self.task_log.append(task)
        task.status = "running"
        log.info("→ Dispatching task %s to %s", task.id[:8], msg.recipient)
        try:
            task.output = await agent.handle(task)
            task.status = "done"
        except Exception as exc:
            task.status = "error"
            task.output = {"error": str(exc)}
            log.error("Task %s failed: %s", task.id[:8], exc)
            raise
        return task

    async def run_pipeline(self):
        """Execute the four-agent pipeline sequentially."""
        tid1 = str(uuid.uuid4())
        await self.send(Message("orchestrator", "EmbeddingAgent", tid1,
                                {"file": HISTORICAL_FILE}))

        tid2 = str(uuid.uuid4())
        await self.send(Message("orchestrator", "AnalysisAgent", tid2,
                                {"file": OPEN_ALERTS_FILE}))

        tid3 = str(uuid.uuid4())
        await self.send(Message("orchestrator", "ReportAgent", tid3,
                                {"open_alerts_file": OPEN_ALERTS_FILE,
                                 "output_file": OUTPUT_FILE}))

        tid4 = str(uuid.uuid4())
        result = await self.send(Message("orchestrator", "ArchiveAgent", tid4,
                                         {"historical_file":  HISTORICAL_FILE,
                                          "report_file":      OUTPUT_FILE,
                                          "kb_processed_dir": str(KB_PROCESSED_DIR)}))
        return result


# ══════════════════════════════════════════════════════════════════════════════
# BASE AGENT
# ══════════════════════════════════════════════════════════════════════════════

class BaseAgent:
    name: str = "BaseAgent"

    def __init__(self):
        self.orchestrator: A2AOrchestrator | None = None

    async def handle(self, task: Task) -> dict:
        raise NotImplementedError

    # ── Embedding helper ───────────────────────────────────────────────────────
    def embed_text(self, text: str) -> list[float]:
        """
        Embeds `text` using a local sentence-transformers model.

        The model runs entirely inside this Python process on CPU — no daemon,
        no HTTP call, no external server.  The SentenceTransformer singleton is
        loaded once on the first call and reused for every subsequent call,
        so the ~1 s model-load cost is paid only once per pipeline run.

        On the very first run the model weights (~22 MB) are downloaded from
        Hugging Face and cached in ~/.cache/huggingface/.  All subsequent runs
        are fully offline.

        Returns a plain Python list[float] compatible with ChromaDB.
        Returns [] on any failure so the caller can skip that alert gracefully.
        """
        global _EMBED_MODEL
        if not RAG_AVAILABLE:
            return []
        try:
            if _EMBED_MODEL is None:
                log.info("[embed] Loading '%s' (first call — "
                         "downloads ~22 MB if not cached)…", EMBED_MODEL_NAME)
                _EMBED_MODEL = SentenceTransformer(EMBED_MODEL_NAME)
                log.info("[embed] Model ready.")
            vector = _EMBED_MODEL.encode(text, show_progress_bar=False)
            return vector.tolist()
        except Exception as exc:
            log.warning("embed_text failed (%s) — skipping this alert", exc)
            return []

    # ── LLM helper ─────────────────────────────────────────────────────────────
    def call_llm(self, prompt: str, max_tokens: int = MAX_TOKENS) -> str:
        payload = {
            "model": LLM_MODEL,
            "prompt": prompt,
            "max_tokens": max_tokens,
            "temperature": TEMPERATURE,
        }
        # Split timeout: (connect, read)
        # A large prompt can take a long time to generate — read timeout is
        # intentionally generous so we don't abort mid-generation.
        timeout = (LLM_TIMEOUT_CONNECT, LLM_TIMEOUT_READ)
        last_exc = None
        for attempt in range(1, LLM_RETRIES + 2):  # 1 original + LLM_RETRIES
            try:
                log.debug("LLM call attempt %d/%d (connect=%ss read=%ss)",
                          attempt, LLM_RETRIES + 1,
                          LLM_TIMEOUT_CONNECT, LLM_TIMEOUT_READ)
                resp = requests.post(LLM_ENDPOINT, json=payload, timeout=timeout)
                resp.raise_for_status()
                data = resp.json()
                # OpenAI-compatible /v1/completions response
                raw_text = data["choices"][0]["text"].strip()
                # Strip markdown code fences the model sometimes wraps around JSON
                if raw_text.startswith("```"):
                    raw_text = raw_text.split("```")[1]
                    if raw_text.lower().startswith("json"):
                        raw_text = raw_text[4:]
                    raw_text = raw_text.strip()
                if not raw_text:
                    log.warning("LLM returned empty response (attempt %d) — "
                                "finish_reason=%s, usage=%s",
                                attempt,
                                data["choices"][0].get("finish_reason", "?"),
                                data.get("usage", {}))
                    raise ValueError("Empty LLM response")
                log.debug("LLM raw response (first 200 chars): %s", raw_text[:200])
                return raw_text
            except requests.exceptions.ConnectTimeout as exc:
                last_exc = exc
                log.warning("LLM attempt %d: connect timeout after %ss — %s",
                            attempt, LLM_TIMEOUT_CONNECT, exc)
            except requests.exceptions.ReadTimeout as exc:
                last_exc = exc
                log.warning("LLM attempt %d: read timeout after %ss — "
                            "model is still generating; consider raising "
                            "LLM_TIMEOUT_READ or reducing prompt size", attempt,
                            LLM_TIMEOUT_READ)
            except requests.exceptions.HTTPError as exc:
                last_exc = exc
                # Only retry on 5xx server errors, not 4xx client errors
                if exc.response is not None and exc.response.status_code < 500:
                    log.error("LLM call failed (client error, no retry): %s", exc)
                    raise
                log.warning("LLM attempt %d: server error %s", attempt, exc)
            except Exception as exc:
                last_exc = exc
                log.error("LLM call failed (non-retryable): %s", exc)
                raise

            if attempt <= LLM_RETRIES:
                wait = attempt * 3  # 3 s, 6 s, …
                log.info("Retrying in %ss…", wait)
                import time; time.sleep(wait)

        log.error("LLM call failed after %d attempts: %s", LLM_RETRIES + 1, last_exc)
        raise last_exc


# ══════════════════════════════════════════════════════════════════════════════
# AGENT 1 — EMBEDDING AGENT  (replaces TrainingAgent)
# ══════════════════════════════════════════════════════════════════════════════

class EmbeddingAgent(BaseAgent):
    """
    RAG Step 1 — Build the vector knowledge base.

    Reads alerts_data.xlsx, converts every row into a dense text representation,
    calls the LLM embedding endpoint to get a vector, and upserts the result into
    a ChromaDB collection persisted on disk.

    Incremental by design: alert IDs already present in the collection are skipped,
    so re-running the pipeline on the same file is cheap (no re-embedding).

    Also stores lightweight in-memory fallback rules (used by AnalysisAgent if
    ChromaDB is unavailable or the embedding call fails).
    """
    name = "EmbeddingAgent"

    @staticmethod
    def _row_to_text(row: dict) -> str:
        """Serialise one alert row into a compact, embedding-friendly string."""
        return (
            f"AlertType={row.get('alert_type','')} "
            f"Score={row.get('score','')} "
            f"Status={row.get('status','')} "
            f"Priority={row.get('priority','')} "
            f"Country={row.get('country','')} "
            f"Currency={row.get('currency','')} "
            f"Amount={row.get('amount','')} "
            f"Description={str(row.get('description',''))[:120]}"
        )

    @staticmethod
    def _row_to_metadata(row: dict) -> dict:
        """Flatten one alert row into ChromaDB-compatible metadata (str/int/float only)."""
        return {
            "sheet":       str(row.get("sheet", "")),
            "alert_id":    str(row.get("alert_id", "")),
            "alert_type":  str(row.get("alert_type", "")),
            "score":       float(row.get("score", 0)),
            "status":      str(row.get("status", "")),
            "priority":    str(row.get("priority", "")),
            "description": str(row.get("description", ""))[:200],
            "country":     str(row.get("country", "")),
            "currency":    str(row.get("currency", "")),
            "amount":      float(row.get("amount", 0)),
        }

    def _get_collection(self):
        """Return (or create) the persisted ChromaDB collection."""
        client = chromadb.PersistentClient(
            path=VECTOR_DB_DIR,
            settings=Settings(anonymized_telemetry=False),
        )
        return client.get_or_create_collection(
            name=COLLECTION_NAME,
            metadata={"hnsw:space": "cosine"},
        )

    async def handle(self, task: Task) -> dict:
        file_path = task.input["file"]

        if not RAG_AVAILABLE:
            log.warning("[EmbeddingAgent] RAG packages not available — "
                        "falling back to rule-based KB. "
                        "Run: pip install chromadb sentence-transformers")
            return self._build_fallback_kb(file_path)

        log.info("[EmbeddingAgent] Loading historical data from %s", file_path)
        all_sheets = pd.read_excel(file_path, sheet_name=None)

        collection   = self._get_collection()
        existing_ids = set(collection.get(include=[])["ids"])
        log.info("[EmbeddingAgent] Collection '%s' has %d existing vectors",
                 COLLECTION_NAME, len(existing_ids))

        embedded = skipped = failed = 0

        for sheet_name, df in all_sheets.items():
            df.columns = [c.strip() for c in df.columns]
            for _, row in df.iterrows():
                alert_id = str(row.get("Alert ID", "")).strip()
                if not alert_id:
                    continue

                # Incremental: skip rows already in the vector store
                if alert_id in existing_ids:
                    skipped += 1
                    continue

                row_dict = {
                    "sheet":       sheet_name,
                    "alert_id":    alert_id,
                    "alert_type":  str(row.get("Alert Type", "")),
                    "alert_type_id": str(row.get("Alert Type ID", "")),
                    "score":       float(row.get("Score", 0) or 0),
                    "status":      str(row.get("Status", "")),
                    "priority":    str(row.get("Priority", "")),
                    "description": str(row.get("Description", "")),
                    "amount":      float(row.get("Amount", 0) or 0),
                    "currency":    str(row.get("Currency", "")),
                    "country":     str(row.get("Country", "")),
                }

                text      = self._row_to_text(row_dict)
                embedding = self.embed_text(text)

                if not embedding:
                    log.warning("[EmbeddingAgent] Skipping %s — empty embedding", alert_id)
                    failed += 1
                    continue

                collection.upsert(
                    ids=[alert_id],
                    embeddings=[embedding],
                    documents=[text],
                    metadatas=[self._row_to_metadata(row_dict)],
                )
                embedded += 1

        total = collection.count()
        log.info("[EmbeddingAgent] Done — embedded=%d skipped=%d failed=%d "
                 "| collection total=%d", embedded, skipped, failed, total)

        # Store collection reference and fallback rules in shared KB
        self.orchestrator.knowledge_base["chroma_collection"] = collection
        self.orchestrator.knowledge_base["rag_available"]     = True
        self._seed_fallback_rules()

        return {
            "status":     "embedded",
            "embedded":   embedded,
            "skipped":    skipped,
            "failed":     failed,
            "total_vectors": total,
        }

    def _seed_fallback_rules(self):
        """Always store lightweight rules as a safety net."""
        self.orchestrator.knowledge_base.setdefault("closure_criteria", [
            "Score < 75 with Low priority → auto-close",
            "Score < 80 with Medium priority and no high-risk country → auto-close",
            "No cross-border component and score < 70 → auto-close",
        ])
        self.orchestrator.knowledge_base.setdefault("high_risk_indicators", [
            "PEP network", "shell company", "high-risk jurisdiction",
            "phantom shipment", "circular transactions",
        ])
        self.orchestrator.knowledge_base.setdefault("auto_close_score_threshold", 75)
        self.orchestrator.knowledge_base.setdefault("never_auto_close_priorities",
                                                    ["Critical", "Escalated"])

    def _build_fallback_kb(self, file_path: str) -> dict:
        """
        Used when chromadb is not installed.
        Reads the file and populates in-memory rules exactly as the old TrainingAgent did,
        so the rest of the pipeline degrades gracefully.
        """
        all_sheets = pd.read_excel(file_path, sheet_name=None)
        patterns: list[dict] = []
        for sheet_name, df in all_sheets.items():
            df.columns = [c.strip() for c in df.columns]
            for _, row in df.iterrows():
                patterns.append({
                    "sheet":       sheet_name,
                    "alert_id":    str(row.get("Alert ID", "")),
                    "alert_type":  str(row.get("Alert Type", "")),
                    "score":       float(row.get("Score", 0) or 0),
                    "status":      str(row.get("Status", "")),
                    "priority":    str(row.get("Priority", "")),
                    "description": str(row.get("Description", "")),
                    "amount":      float(row.get("Amount", 0) or 0),
                    "currency":    str(row.get("Currency", "")),
                    "country":     str(row.get("Country", "")),
                })
        kb = {
            "closure_criteria": [
                "Score < 75 with Low priority → auto-close",
                "Score < 80 with Medium priority and no high-risk country → auto-close",
            ],
            "high_risk_indicators": ["PEP network", "shell company",
                                     "high-risk jurisdiction", "phantom shipment"],
            "auto_close_score_threshold": 75,
            "never_auto_close_priorities": ["Critical", "Escalated"],
            "all_historical_patterns": patterns,
            "rag_available": False,
        }
        self.orchestrator.knowledge_base.update(kb)
        log.info("[EmbeddingAgent] Fallback KB built with %d patterns.", len(patterns))
        return {"status": "fallback_kb", "patterns_count": len(patterns)}


# ══════════════════════════════════════════════════════════════════════════════
# AGENT 2 — ANALYSIS AGENT
# ══════════════════════════════════════════════════════════════════════════════

class AnalysisAgent(BaseAgent):
    """
    Reads open_alerts_data.xlsx and for each open alert calls the LLM
    (with the knowledge base as context) to decide: auto-close or escalate,
    and generates a closure comment.
    Stores results in the shared knowledge base.
    """
    name = "AnalysisAgent"

    async def handle(self, task: Task) -> dict:
        file_path = task.input["file"]
        kb = self.orchestrator.knowledge_base

        if not Path(file_path).exists():
            raise FileNotFoundError(
                f"{file_path} not found. Please place it in the working directory."
            )

        log.info("[AnalysisAgent] Loading open alerts from %s", file_path)
        all_sheets = pd.read_excel(file_path, sheet_name=None)

        criteria_text = "\n".join(
            f"  {i+1}. {c}" for i, c in enumerate(kb.get("closure_criteria", []))
        )
        high_risk_text = ", ".join(kb.get("high_risk_indicators", []))
        threshold = kb.get("auto_close_score_threshold", 75)
        never_close = kb.get("never_auto_close_priorities", ["Critical"])

        decisions: list[dict] = []

        for sheet_name, df in all_sheets.items():
            df.columns = [c.strip() for c in df.columns]
            open_alerts = df[df["Status"].str.strip().str.lower() == "open"]
            log.info("[AnalysisAgent] Sheet '%s': %d open alerts", sheet_name, len(open_alerts))

            for _, row in open_alerts.iterrows():
                alert = {
                    "sheet": sheet_name,
                    "alert_id": str(row.get("Alert ID", "")),
                    "customer_id": str(row.get("Customer ID", "")),
                    "customer_name": str(row.get("Customer Name", "")),
                    "alert_type": str(row.get("Alert Type", "")),
                    "alert_type_id": str(row.get("Alert Type ID", "")),
                    "score": float(row.get("Score", 0)),
                    "priority": str(row.get("Priority", "")),
                    "description": str(row.get("Description", "")),
                    "amount": float(row.get("Amount", 0)),
                    "currency": str(row.get("Currency", "")),
                    "country": str(row.get("Country", "")),
                    "created_date": str(row.get("Created Date", "")),
                }

                decision = self._analyse_alert(alert, criteria_text,
                                               high_risk_text, threshold, never_close)
                decisions.append(decision)
                log.info("[AnalysisAgent] %s → %s (confidence=%s)",
                         alert["alert_id"], decision["action"], decision["confidence"])

        self.orchestrator.knowledge_base["decisions"] = decisions
        log.info("[AnalysisAgent] Analysis complete. %d decisions made.", len(decisions))
        return {"status": "analysed", "decisions_count": len(decisions)}

    # Compact prompt sent per-alert — kept short to avoid truncation
    _ALERT_PROMPT = (
        "You are an AML compliance officer. Analyse this alert and respond with "
        "ONLY a JSON object — no explanation, no markdown, no extra text.\n\n"
        "Rules:\n"
        "  Score threshold for auto-closure: {threshold}\n"
        "  Never auto-close priorities: {never_close}\n"
        "  High-risk indicators (escalate if present): {high_risk}\n\n"
        "Alert:\n"
        "  ID={alert_id} Type={alert_type} Score={score} "
        "Priority={priority} Country={country}\n"
        "  Amount={amount} {currency}\n"
        "  Description: {description}\n\n"
        "Required JSON keys (respond with NOTHING else):\n"
        '  {{"action":"AUTO_CLOSE|ESCALATE|REVIEW",'
        '"confidence":"HIGH|MEDIUM|LOW",'
        '"comment":"<2 sentence professional closure comment>",'
        '"risk_flags":["flag1","flag2"]}}'
    )

    def _retrieve_similar(self, alert: dict) -> str:
        """
        RAG retrieval step.
        Embeds the open alert and queries ChromaDB for RAG_TOP_K most similar
        historical alerts. Returns a formatted context string for the LLM prompt.
        Returns empty string if RAG is unavailable.
        """
        kb         = self.orchestrator.knowledge_base
        collection = kb.get("chroma_collection")
        if collection is None:
            return ""

        query_text = (
            f"AlertType={alert.get('alert_type','')} "
            f"Score={alert.get('score','')} "
            f"Status=Open "
            f"Priority={alert.get('priority','')} "
            f"Country={alert.get('country','')} "
            f"Currency={alert.get('currency','')} "
            f"Amount={alert.get('amount','')} "
            f"Description={str(alert.get('description',''))[:120]}"
        )

        embedding = self.embed_text(query_text)
        if not embedding:
            log.warning("[AnalysisAgent] Could not embed alert %s — skipping RAG",
                        alert["alert_id"])
            return ""

        try:
            results = collection.query(
                query_embeddings=[embedding],
                n_results=min(RAG_TOP_K, collection.count()),
                include=["documents", "metadatas", "distances"],
            )
        except Exception as exc:
            log.warning("[AnalysisAgent] ChromaDB query failed for %s: %s",
                        alert["alert_id"], exc)
            return ""

        metadatas = results.get("metadatas", [[]])[0]
        distances = results.get("distances", [[]])[0]

        if not metadatas:
            return ""

        lines = [f"Top-{len(metadatas)} similar historical alerts (cosine distance):"]
        for i, (meta, dist) in enumerate(zip(metadatas, distances), 1):
            similarity = round((1 - dist) * 100, 1)
            lines.append(
                f"  {i}. [{meta.get('alert_id','')}] "
                f"Type={meta.get('alert_type','')} "
                f"Score={meta.get('score','')} "
                f"Status={meta.get('status','')} "
                f"Priority={meta.get('priority','')} "
                f"Country={meta.get('country','')} "
                f"Desc={meta.get('description','')[:80]} "
                f"(similarity={similarity}%)"
            )

        log.debug("[AnalysisAgent] RAG retrieved %d examples for %s",
                  len(metadatas), alert["alert_id"])
        return "\n".join(lines)

    # RAG-enhanced prompt — similar examples replace the full KB dump
    _RAG_PROMPT = (
        "You are an AML compliance officer. Use the similar historical cases below "
        "to decide how to handle the new alert. Respond with ONLY a JSON object — "
        "no explanation, no markdown, no extra text.\n\n"
        "Historical similar cases (learn the pattern from Status field):\n"
        "{similar_cases}\n\n"
        "Rules:\n"
        "  Score threshold for auto-closure: {threshold}\n"
        "  Never auto-close priorities: {never_close}\n"
        "  High-risk indicators (escalate if present): {high_risk}\n\n"
        "New alert to decide:\n"
        "  ID={alert_id} Type={alert_type} Score={score} "
        "Priority={priority} Country={country}\n"
        "  Amount={amount} {currency}\n"
        "  Description: {description}\n\n"
        'Required JSON: {{"action":"AUTO_CLOSE|ESCALATE|REVIEW",'
        '"confidence":"HIGH|MEDIUM|LOW",'
        '"comment":"<2 sentence professional comment>",'
        '"risk_flags":["flag1"]}}'
    )

    def _analyse_alert(self, alert: dict, criteria_text: str,
                       high_risk_text: str, threshold: int,
                       never_close: list) -> dict:
        # ── RAG retrieval ─────────────────────────────────────────────────────
        similar_cases = self._retrieve_similar(alert)
        rag_used      = bool(similar_cases)

        if rag_used:
            prompt = self._RAG_PROMPT.format(
                similar_cases = similar_cases,
                threshold     = threshold,
                never_close   = ", ".join(never_close),
                high_risk     = high_risk_text[:200],
                alert_id      = alert["alert_id"],
                alert_type    = alert["alert_type"],
                score         = alert["score"],
                priority      = alert["priority"],
                country       = alert["country"],
                amount        = alert["amount"],
                currency      = alert["currency"],
                description   = alert["description"][:120],
            )
        else:
            # Fallback: compact rule-based prompt (no similar cases available)
            prompt = self._ALERT_PROMPT.format(
                threshold    = threshold,
                never_close  = ", ".join(never_close),
                high_risk    = high_risk_text[:200],
                alert_id     = alert["alert_id"],
                alert_type   = alert["alert_type"],
                score        = alert["score"],
                priority     = alert["priority"],
                country      = alert["country"],
                amount       = alert["amount"],
                currency     = alert["currency"],
                description  = alert["description"][:120],
            )

        result = None
        try:
            raw    = self.call_llm(prompt, max_tokens=200)
            result = self._extract_json(raw, alert["alert_id"])
        except Exception as exc:
            log.warning("[AnalysisAgent] LLM parse error for %s: %s — using rule-based fallback",
                        alert["alert_id"], exc)

        if result is None:
            result = self._rule_based_decision(alert, threshold, never_close, high_risk_text)

        # Normalise: guarantee all required keys exist
        result.setdefault("action",     "REVIEW")
        result.setdefault("confidence", "LOW")
        result.setdefault("comment",    "Decision made via fallback rule engine.")
        result.setdefault("risk_flags", [])
        result["action"]   = str(result["action"]).upper()
        result["rag_used"] = rag_used          # trace flag for the report

        result["alert"] = alert
        return result

    @staticmethod
    def _extract_json(raw: str, alert_id: str) -> dict:
        """
        Robustly extract a JSON object from the LLM response.
        Handles: leading/trailing text, markdown fences, single-quoted keys,
        partial responses that still contain a valid { ... } block.
        """
        if not raw:
            raise ValueError("Empty response from LLM")

        # 1. Try parsing the whole string first (ideal case)
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            pass

        # 2. Find the first '{' and last '}' and try just that substring
        start = raw.find("{")
        end   = raw.rfind("}")
        if start != -1 and end != -1 and end > start:
            candidate = raw[start:end + 1]
            try:
                return json.loads(candidate)
            except json.JSONDecodeError:
                pass

            # 3. Replace Python/JS-style single quotes with double quotes
            #    and try once more (some models output {'key':'val'})
            import re
            fixed = re.sub(r"(?<![\\])'", '"', candidate)
            try:
                return json.loads(fixed)
            except json.JSONDecodeError:
                pass

        log.warning("[AnalysisAgent] Could not extract JSON for %s. Raw (200 chars): %s",
                    alert_id, raw[:200])
        return None

    def _rule_based_decision(self, alert: dict, threshold: int,
                              never_close: list, high_risk_text: str) -> dict:
        """Deterministic fallback if LLM call fails."""
        score = alert["score"]
        priority = alert["priority"]
        desc = alert["description"].lower()
        risk_flags = [kw for kw in high_risk_text.split(", ") if kw.lower() in desc]

        if priority in never_close or risk_flags:
            return {
                "action": "ESCALATE",
                "confidence": "HIGH",
                "comment": (
                    f"Alert {alert['alert_id']} has been flagged for escalation due to "
                    f"{'high priority level' if priority in never_close else 'presence of high-risk indicators'}. "
                    "Manual review by a senior analyst is required before any closure decision."
                ),
                "risk_flags": risk_flags,
            }
        elif score <= threshold:
            return {
                "action": "AUTO_CLOSE",
                "confidence": "MEDIUM",
                "comment": (
                    f"Alert {alert['alert_id']} meets the auto-closure criteria with a risk score "
                    f"of {score} (below threshold of {threshold}) and {priority} priority. "
                    "No high-risk indicators detected in the transaction description."
                ),
                "risk_flags": [],
            }
        else:
            return {
                "action": "REVIEW",
                "confidence": "LOW",
                "comment": (
                    f"Alert {alert['alert_id']} (score {score}) requires further review. "
                    "Risk score exceeds auto-closure threshold but no immediate escalation triggers found. "
                    "Assign to analyst for manual assessment."
                ),
                "risk_flags": [],
            }


# ══════════════════════════════════════════════════════════════════════════════
# AGENT 3 — REPORT AGENT
# ══════════════════════════════════════════════════════════════════════════════

class ReportAgent(BaseAgent):
    """
    Generates the auto-closure report Excel and writes decisions + comments
    back into the open_alerts_data.xlsx file.
    """
    name = "ReportAgent"

    # colour palette
    COLOURS = {
        "AUTO_CLOSE": "C6EFCE",   # green
        "ESCALATE":   "FFCCCC",   # red
        "REVIEW":     "FFEB9C",   # amber
        "header_bg":  "1F4E79",   # dark blue
        "header_fg":  "FFFFFF",
        "sub_bg":     "D6E4F0",
    }

    async def handle(self, task: Task) -> dict:
        open_file  = task.input["open_alerts_file"]
        output     = task.input["output_file"]
        decisions  = self.orchestrator.knowledge_base.get("decisions", [])
        kb         = self.orchestrator.knowledge_base

        log.info("[ReportAgent] Building auto-closure report for %d decisions", len(decisions))

        # 1. Write decisions back into open_alerts_data.xlsx
        self._annotate_source_file(open_file, decisions)

        # 2. Build standalone report workbook
        self._build_report_workbook(output, decisions, kb)

        auto_closed = sum(1 for d in decisions if d["action"] == "AUTO_CLOSE")
        escalated   = sum(1 for d in decisions if d["action"] == "ESCALATE")
        review      = sum(1 for d in decisions if d["action"] == "REVIEW")

        log.info("[ReportAgent] Report saved to %s — AUTO_CLOSE=%d ESCALATE=%d REVIEW=%d",
                 output, auto_closed, escalated, review)
        return {"output_file": output, "auto_closed": auto_closed,
                "escalated": escalated, "review": review}

    # ── annotate source file ────────────────────────────────────────────────────

    def _annotate_source_file(self, file_path: str, decisions: list[dict]):
        """Add Action, Comment, Risk Flags, Processed At columns to open_alerts_data.xlsx."""
        if not Path(file_path).exists():
            log.warning("[ReportAgent] %s not found; skipping annotation.", file_path)
            return

        wb = load_workbook(file_path)
        decision_map = {d["alert"]["alert_id"]: d for d in decisions}

        for ws in wb.worksheets:
            headers = [cell.value for cell in ws[1]]
            if "Alert ID" not in headers:
                continue

            id_col = headers.index("Alert ID") + 1   # 1-based

            # Add new columns if not present
            new_cols = ["Action", "Closure Comment", "Risk Flags", "Processed At"]
            for col_name in new_cols:
                if col_name not in headers:
                    ws.cell(row=1, column=len(headers) + 1,
                            value=col_name).font = Font(bold=True)
                    headers.append(col_name)

            action_col  = headers.index("Action") + 1
            comment_col = headers.index("Closure Comment") + 1
            flags_col   = headers.index("Risk Flags") + 1
            ts_col      = headers.index("Processed At") + 1
            status_col  = headers.index("Status") + 1 if "Status" in headers else None

            for row_idx in range(2, ws.max_row + 1):
                alert_id = str(ws.cell(row=row_idx, column=id_col).value or "")
                if alert_id not in decision_map:
                    continue
                dec = decision_map[alert_id]
                action = dec.get("action", "")
                fill = PatternFill("solid",
                                   start_color=self.COLOURS.get(action, "FFFFFF"))

                ws.cell(row=row_idx, column=action_col, value=action).fill = fill
                ws.cell(row=row_idx, column=comment_col,
                        value=dec.get("comment", "")).alignment = Alignment(wrap_text=True)
                ws.cell(row=row_idx, column=flags_col,
                        value=", ".join(dec.get("risk_flags", [])))
                ws.cell(row=row_idx, column=ts_col,
                        value=datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"))

                if status_col and action == "AUTO_CLOSE":
                    ws.cell(row=row_idx, column=status_col, value="Closed")

        wb.save(file_path)
        log.info("[ReportAgent] Annotated %s", file_path)

    # ── standalone report workbook ──────────────────────────────────────────────

    def _build_report_workbook(self, output: str, decisions: list[dict], kb: dict):
        from openpyxl import Workbook

        wb = Workbook()
        self._write_summary_sheet(wb.active, decisions, kb)
        self._write_detail_sheet(wb.create_sheet("Alert Decisions"), decisions)
        self._write_kb_sheet(wb.create_sheet("Knowledge Base"), kb)
        wb.save(output)

    def _hdr(self, cell, text: str, bold=True, bg=None, fg=None):
        cell.value = text
        cell.font = Font(bold=bold,
                         color=fg or self.COLOURS["header_fg"],
                         name="Arial", size=11)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_summary_sheet(self, ws, decisions, kb):
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

        auto   = [d for d in decisions if d["action"] == "AUTO_CLOSE"]
        esc    = [d for d in decisions if d["action"] == "ESCALATE"]
        review = [d for d in decisions if d["action"] == "REVIEW"]

        ws.row_dimensions[1].height = 36
        self._hdr(ws["A1"], "AML ALERT AUTO-CLOSURE REPORT", bg=self.COLOURS["header_bg"])
        ws.merge_cells("A1:B1")

        rows = [
            ("Generated At",    datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Open Alerts Analysed", len(decisions)),
            ("Auto-Closed",     len(auto)),
            ("Escalated",       len(esc)),
            ("Needs Review",    len(review)),
            ("Score Threshold", kb.get("auto_close_score_threshold", "N/A")),
        ]
        for r, (k, v) in enumerate(rows, start=2):
            ws.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
            ws.cell(row=r, column=2, value=v).alignment = Alignment(horizontal="center")

        # Mini table: action breakdown by sheet
        row = len(rows) + 3
        self._hdr(ws.cell(row=row, column=1), "Sheet", bg=self.COLOURS["sub_bg"],
                  fg="000000")
        self._hdr(ws.cell(row=row, column=2), "# Alerts", bg=self.COLOURS["sub_bg"],
                  fg="000000")
        row += 1
        sheets_seen = {}
        for d in decisions:
            s = d["alert"].get("sheet", "Unknown")
            sheets_seen[s] = sheets_seen.get(s, 0) + 1
        for sheet_name, cnt in sheets_seen.items():
            ws.cell(row=row, column=1, value=sheet_name)
            ws.cell(row=row, column=2, value=cnt)
            row += 1

    def _write_detail_sheet(self, ws, decisions):
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 50
        ws.column_dimensions["J"].width = 30

        headers = ["Alert ID", "Sheet", "Customer Name", "Alert Type",
                   "Score", "Priority", "Action", "Confidence",
                   "Closure Comment", "Risk Flags"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color=self.COLOURS["header_fg"],
                             name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=self.COLOURS["header_bg"])
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 22

        for r, d in enumerate(decisions, start=2):
            a = d["alert"]
            action = d.get("action", "")
            fill = PatternFill("solid",
                               start_color=self.COLOURS.get(action, "FFFFFF"))
            row_data = [
                a.get("alert_id", ""),
                a.get("sheet", ""),
                a.get("customer_name", ""),
                a.get("alert_type", ""),
                a.get("score", ""),
                a.get("priority", ""),
                action,
                d.get("confidence", ""),
                d.get("comment", ""),
                ", ".join(d.get("risk_flags", [])),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                if col in (7, 8):   # Action, Confidence columns get colour
                    cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 48

    def _write_kb_sheet(self, ws, kb):
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 80

        self._hdr(ws.cell(row=1, column=1), "#",
                  bg=self.COLOURS["header_bg"])
        self._hdr(ws.cell(row=1, column=2), "Closure Criteria (learned from historical data)",
                  bg=self.COLOURS["header_bg"])

        for i, rule in enumerate(kb.get("closure_criteria", []), start=2):
            ws.cell(row=i, column=1, value=i - 1)
            cell = ws.cell(row=i, column=2, value=rule)
            cell.alignment = Alignment(wrap_text=True)

        row = len(kb.get("closure_criteria", [])) + 3
        ws.cell(row=row, column=1,
                value="High-Risk Indicators").font = Font(bold=True)
        ws.cell(row=row, column=2,
                value=", ".join(kb.get("high_risk_indicators", [])))



# ══════════════════════════════════════════════════════════════════════════════
# AGENT 4 — ARCHIVE AGENT
# ══════════════════════════════════════════════════════════════════════════════

class ArchiveAgent(BaseAgent):
    """
    Runs after ReportAgent. Responsibilities in order:

    1. ARCHIVE  — copy alerts_data.xlsx into KB_Processed/ with a timestamped
                  name, then merge it into a single cumulative master file
                  (KB_Processed/master_kb.xlsx).  Every run appends; the master
                  file is created on the first run automatically.

    2. PROMOTE  — rename auto_closure_report.xlsx → alerts_data.xlsx so the
                  next pipeline run trains on the freshly decided data.
    """
    name = "ArchiveAgent"

    # ── Column sets we expect across sheets ────────────────────────────────────
    # The master KB inherits whatever columns exist; new columns from later runs
    # are appended with NaN for older rows — pandas handles this automatically.

    async def handle(self, task: Task) -> dict:
        historical  = Path(task.input["historical_file"])
        report      = Path(task.input["report_file"])
        kb_dir      = Path(task.input["kb_processed_dir"])
        run_ts      = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

        # ── Validate prerequisites ──────────────────────────────────────────────
        if not historical.exists():
            raise FileNotFoundError(f"[ArchiveAgent] Historical file missing: {historical}")
        if not report.exists():
            raise FileNotFoundError(f"[ArchiveAgent] Report file missing: {report}")

        # ── Step 1: Create KB_Processed/ if it doesn't exist ───────────────────
        kb_dir.mkdir(parents=True, exist_ok=True)
        log.info("[ArchiveAgent] KB_Processed dir: %s", kb_dir)

        # ── Step 2: Copy alerts_data.xlsx → KB_Processed/<timestamp>_alerts_data.xlsx
        archived_name = f"{run_ts}_alerts_data.xlsx"
        archived_path = kb_dir / archived_name
        import shutil
        shutil.copy2(historical, archived_path)
        log.info("[ArchiveAgent] Archived → %s", archived_path)

        # ── Step 3: Merge archived file into master_kb.xlsx ────────────────────
        master_path   = kb_dir / "master_kb.xlsx"
        merge_result  = self._merge_into_master(archived_path, master_path, run_ts)

        # ── Step 4: Promote report → alerts_data.xlsx ──────────────────────────
        # Remove the old historical file first, then move the report into its place.
        historical.unlink()
        shutil.move(str(report), str(historical))
        log.info("[ArchiveAgent] Promoted %s → %s", report.name, historical.name)

        return {
            "archived_as":       archived_name,
            "master_kb":         str(master_path),
            "master_total_rows": merge_result["total_rows"],
            "master_run_count":  merge_result["run_count"],
            "new_alerts_data":   str(historical),
        }

    # ── Merge helper ────────────────────────────────────────────────────────────

    def _merge_into_master(self, source: Path, master: Path, run_ts: str) -> dict:
        """
        Reads `source` (the just-archived snapshot) and appends every sheet's
        rows into the matching sheet of `master`.  A '_run_timestamp' column is
        added to every row so you can trace which pipeline run produced it.

        Sheet names are matched by name; if a sheet appears in `source` but not
        yet in `master`, it is created.  If `master` does not exist yet (first
        run), it is created from scratch.
        """
        # Read incoming snapshot
        source_sheets: dict[str, pd.DataFrame] = pd.read_excel(source, sheet_name=None)

        # Stamp every row with the run timestamp
        for sheet_name, df in source_sheets.items():
            df["_run_timestamp"] = run_ts
            source_sheets[sheet_name] = df

        if not master.exists():
            # ── First run: just write the source as-is ─────────────────────────
            log.info("[ArchiveAgent] master_kb.xlsx does not exist — creating fresh.")
            with pd.ExcelWriter(str(master), engine="openpyxl") as writer:
                for sheet_name, df in source_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            total_rows = sum(len(df) for df in source_sheets.values())
            self._apply_master_formatting(master)
            log.info("[ArchiveAgent] master_kb.xlsx created with %d rows.", total_rows)
            return {"total_rows": total_rows, "run_count": 1}

        # ── Subsequent runs: load existing master and append ───────────────────
        existing_sheets: dict[str, pd.DataFrame] = pd.read_excel(master, sheet_name=None)

        # Detect run count from existing data
        sample_df = next(iter(existing_sheets.values()), pd.DataFrame())
        if "_run_timestamp" in sample_df.columns:
            run_count = sample_df["_run_timestamp"].nunique() + 1
        else:
            run_count = 2   # master existed but had no timestamp column

        merged_sheets: dict[str, pd.DataFrame] = {}

        all_sheet_names = set(existing_sheets) | set(source_sheets)
        for sheet_name in all_sheet_names:
            existing_df = existing_sheets.get(sheet_name, pd.DataFrame())
            incoming_df = source_sheets.get(sheet_name, pd.DataFrame())
            merged = pd.concat([existing_df, incoming_df], ignore_index=True, sort=False)
            merged_sheets[sheet_name] = merged

        total_rows = sum(len(df) for df in merged_sheets.values())

        # Write merged master (overwrite in place)
        with pd.ExcelWriter(str(master), engine="openpyxl") as writer:
            for sheet_name, df in merged_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        self._apply_master_formatting(master)
        log.info("[ArchiveAgent] master_kb.xlsx updated — run #%d, %d total rows.",
                 run_count, total_rows)
        return {"total_rows": total_rows, "run_count": run_count}

    # ── Light formatting for the master file ────────────────────────────────────

    def _apply_master_formatting(self, master: Path):
        """Freeze header row, bold headers, auto-filter on every sheet."""
        from openpyxl.utils import get_column_letter
        wb = load_workbook(str(master))
        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue
            ws.freeze_panes = "A2"
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
            for cell in ws[1]:
                cell.font = Font(bold=True, name="Calibri", size=10,
                                 color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Highlight the _run_timestamp column in a muted teal so it stands out
            for col_idx, cell in enumerate(ws[1], 1):
                if str(cell.value) == "_run_timestamp":
                    for row_idx in range(1, ws.max_row + 1):
                        rc = ws.cell(row=row_idx, column=col_idx)
                        rc.fill = PatternFill("solid", start_color="D6E4F0")
        wb.save(str(master))


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

async def main():
    print("\n" + "═" * 70)
    print("  AML Alert Auto-Closure System — A2A Agent Pipeline")
    print("═" * 70 + "\n")

    orchestrator = A2AOrchestrator()
    orchestrator.register(EmbeddingAgent())
    orchestrator.register(AnalysisAgent())
    orchestrator.register(ReportAgent())
    orchestrator.register(ArchiveAgent())

    result = await orchestrator.run_pipeline()
    rag_on = orchestrator.knowledge_base.get("rag_available", False)
    print("\n" + "═" * 70)
    print(f"  Pipeline Complete!  [RAG={'ON' if rag_on else 'OFF (fallback)'}]")
    print(f"  ✓ Auto-Closed  : {result.output.get('auto_closed', 0)}")
    print(f"  ✗ Escalated    : {result.output.get('escalated', 0)}")
    print(f"  ? Needs Review : {result.output.get('review', 0)}")
    print(f"  Archived as    : {result.output.get('archived_as', '—')}")
    print(f"  Master KB rows : {result.output.get('master_total_rows', '—')}")
    print(f"  Master KB runs : {result.output.get('master_run_count', '—')}")
    print(f"  New training   : {result.output.get('new_alerts_data', '—')}")
    print("═" * 70 + "\n")

    print("\nFull task log:")
    for t in orchestrator.task_log:
        print(f"  [{t.status.upper():7s}] {t.name:20s} — {json.dumps(t.output)}")


if __name__ == "__main__":
    asyncio.run(main())
