"""
agent_report.py
===============
Agent 3 — ReportAgent

Takes the decisions produced by AnalysisAgent and builds a standalone
date-stamped report in the "Closure Report" folder:

    Closure Report/<YYYYMMDD_HHMMSS>_auto_closure_report.xlsx

The report contains three sheets:
    Summary        — KPI counts and per-sheet breakdown
    Alert Decisions— full colour-coded decision table with all open alert
                     fields plus the model's Action, Confidence, Comment,
                     Risk Flags, and RAG flag
    Knowledge Base — closure criteria seeded from historical data

open_alerts_data.xlsx in "Open Alerts" is NEVER modified.
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config as cfg
from orchestrator import BaseAgent, Task

log = logging.getLogger("AML-A2A")


class ReportAgent(BaseAgent):
    name = "ReportAgent"

    COLOURS = {
        "AUTO_CLOSE": "C6EFCE",   # green
        "ESCALATE":   "FFCCCC",   # red
        "REVIEW":     "FFEB9C",   # amber
        "header_bg":  "1F4E79",   # dark navy
        "header_fg":  "FFFFFF",
        "sub_bg":     "D6E4F0",
    }

    # ── Public entry point ─────────────────────────────────────────────────────

    async def handle(self, task: Task) -> dict:
        # open_alerts_file is read-only — we only read it to copy original
        # alert fields into the report. We never write back to it.
        open_file = task.input["open_alerts_file"]
        output    = task.input["output_file"]
        decisions = self.orchestrator.knowledge_base.get("decisions", [])
        kb        = self.orchestrator.knowledge_base

        log.info("[ReportAgent] Building report for %d decisions → %s",
                 len(decisions), Path(output).name)

        self._build_report_workbook(output, decisions, kb)

        auto_closed = sum(1 for d in decisions if d["action"] == "AUTO_CLOSE")
        escalated   = sum(1 for d in decisions if d["action"] == "ESCALATE")
        review      = sum(1 for d in decisions if d["action"] == "REVIEW")

        log.info("[ReportAgent] Saved %s — AUTO_CLOSE=%d ESCALATE=%d REVIEW=%d",
                 Path(output).name, auto_closed, escalated, review)
        return {
            "output_file": output,
            "auto_closed": auto_closed,
            "escalated":   escalated,
            "review":      review,
        }

    # ── Build standalone report workbook ───────────────────────────────────────

    def _build_report_workbook(self, output: str,
                               decisions: list[dict], kb: dict):
        wb = Workbook()
        self._write_summary_sheet(wb.active, decisions, kb)
        self._write_detail_sheet(wb.create_sheet("Alert Decisions"), decisions)
        self._write_kb_sheet(wb.create_sheet("Knowledge Base"), kb)
        wb.save(output)

    # ── Sheet: Summary ─────────────────────────────────────────────────────────

    def _write_summary_sheet(self, ws, decisions, kb):
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

        auto   = [d for d in decisions if d["action"] == "AUTO_CLOSE"]
        esc    = [d for d in decisions if d["action"] == "ESCALATE"]
        review = [d for d in decisions if d["action"] == "REVIEW"]
        rag_ct = sum(1 for d in decisions if d.get("rag_used"))

        ws.row_dimensions[1].height = 36
        self._hdr(ws["A1"], "AML ALERT AUTO-CLOSURE REPORT",
                  bg=self.COLOURS["header_bg"])
        ws.merge_cells("A1:B1")

        rows = [
            ("Generated At",            datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Alerts Analysed",   len(decisions)),
            ("Auto-Closed",             len(auto)),
            ("Escalated",               len(esc)),
            ("Needs Review",            len(review)),
            ("RAG-assisted decisions",  rag_ct),
            ("Score Threshold",         kb.get("auto_close_score_threshold", "N/A")),
        ]
        for r, (k, v) in enumerate(rows, start=2):
            ws.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
            ws.cell(row=r, column=2, value=v).alignment = Alignment(horizontal="center")

        # Per-sheet breakdown
        row = len(rows) + 3
        self._hdr(ws.cell(row=row, column=1), "Sheet",
                  bg=self.COLOURS["sub_bg"], fg="000000")
        self._hdr(ws.cell(row=row, column=2), "# Alerts",
                  bg=self.COLOURS["sub_bg"], fg="000000")
        row += 1
        sheets_seen: dict[str, int] = {}
        for d in decisions:
            s = d["alert"].get("sheet", "Unknown")
            sheets_seen[s] = sheets_seen.get(s, 0) + 1
        for sheet_name, cnt in sheets_seen.items():
            ws.cell(row=row, column=1, value=sheet_name)
            ws.cell(row=row, column=2, value=cnt)
            row += 1

    # ── Sheet: Alert Decisions ─────────────────────────────────────────────────

    def _write_detail_sheet(self, ws, decisions):
        col_widths = [12, 14, 20, 24, 10, 12, 12, 14, 6, 50, 30]
        for i, w in enumerate(col_widths, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w

        headers = ["Alert ID", "Sheet", "Customer Name", "Alert Type",
                   "Score", "Priority", "Action", "Confidence", "RAG",
                   "Closure Comment", "Risk Flags"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color=self.COLOURS["header_fg"],
                                  name="Arial", size=10)
            cell.fill      = PatternFill("solid",
                                         start_color=self.COLOURS["header_bg"])
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 22

        for r, d in enumerate(decisions, start=2):
            a      = d["alert"]
            action = d.get("action", "")
            fill   = PatternFill("solid",
                                 start_color=self.COLOURS.get(action, "FFFFFF"))
            row_data = [
                a.get("alert_id", ""),
                a.get("sheet", ""),
                a.get("customer_name", ""),
                a.get("alert_type", ""),
                a.get("score", ""),
                a.get("priority", ""),
                action,
                d.get("confidence", ""),
                "✓" if d.get("rag_used") else "—",
                d.get("comment", ""),
                ", ".join(d.get("risk_flags", [])),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col in (7, 8):   # Action and Confidence get colour fill
                    cell.fill = fill
            ws.row_dimensions[r].height = 48

    # ── Sheet: Knowledge Base ──────────────────────────────────────────────────

    def _write_kb_sheet(self, ws, kb):
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 80

        self._hdr(ws.cell(row=1, column=1), "#",
                  bg=self.COLOURS["header_bg"])
        self._hdr(ws.cell(row=1, column=2),
                  "Closure Criteria (seeded from historical data)",
                  bg=self.COLOURS["header_bg"])

        for i, rule in enumerate(kb.get("closure_criteria", []), start=2):
            ws.cell(row=i, column=1, value=i - 1)
            ws.cell(row=i, column=2, value=rule).alignment = Alignment(wrap_text=True)

        row = len(kb.get("closure_criteria", [])) + 3
        ws.cell(row=row, column=1, value="High-Risk Indicators").font = Font(bold=True)
        ws.cell(row=row, column=2,
                value=", ".join(kb.get("high_risk_indicators", [])))

    # ── Shared header helper ───────────────────────────────────────────────────

    def _hdr(self, cell, text: str, bold=True, bg=None, fg=None):
        cell.value     = text
        cell.font      = Font(bold=bold,
                              color=fg or self.COLOURS["header_fg"],
                              name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
